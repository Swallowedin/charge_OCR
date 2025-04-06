import streamlit as st
import os
import pandas as pd
import numpy as np
import pytesseract
import cv2
from PIL import Image
import json
import csv
from openai import OpenAI
import tempfile
import io
from pathlib import Path
import base64
import re

# Configuration du logging
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("charges_analyzer.log"),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger("charges_analyzer")

# CSS personnalisé
st.markdown("""
<style>
    .main {
        padding: 1rem;
    }
    .reportview-container .main .block-container {
        max-width: 1000px;
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
</style>
""", unsafe_allow_html=True)

class ChargesAnalyzer:
    """
    Classe pour analyser les tableaux de charges des baux commerciaux 
    à partir de différents formats (Excel, CSV, PDF) en utilisant l'OCR et GPT-4o-mini.
    """
    
    def __init__(self, api_key=None):
        """
        Initialise l'analyseur de charges.
        
        Args:
            api_key (str, optional): Clé API OpenAI.
        """
        self.api_key = api_key
        if not self.api_key:
            raise ValueError("Clé API OpenAI non fournie.")
            
        # Initialiser le client OpenAI
        self.client = OpenAI(api_key=self.api_key)
        
        # Configuration de Tesseract OCR
        self.tesseract_config = r'--oem 3 --psm 6'
    
    def process_file(self, uploaded_file):
        """
        Traite un fichier téléchargé par l'utilisateur.
        
        Args:
            uploaded_file (UploadedFile): Fichier téléchargé via Streamlit.
            
        Returns:
            dict: Résultats de l'analyse structurés.
        """
        file_extension = Path(uploaded_file.name).suffix.lower()
        
        # Créer un fichier temporaire
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
            # Écrire le contenu du fichier téléchargé dans le fichier temporaire
            tmp_file.write(uploaded_file.getvalue())
            tmp_path = tmp_file.name
        
        try:
            if file_extension == '.pdf':
                result = self.analyze_pdf(tmp_path)
            elif file_extension in ['.xlsx', '.xls']:
                result = self.analyze_excel(tmp_path)
            elif file_extension == '.csv':
                result = self.analyze_csv(tmp_path)
            else:
                raise ValueError(f"Format de fichier non supporté: {file_extension}")
            
            # Supprimer le fichier temporaire
            os.unlink(tmp_path)
            
            return result
        except Exception as e:
            # Supprimer le fichier temporaire en cas d'erreur
            os.unlink(tmp_path)
            raise e
    
    def analyze_pdf(self, pdf_path):
        """
        Analyse un fichier PDF en utilisant l'OCR et l'extraction directe ou GPT-4o-mini.
        
        Args:
            pdf_path (str): Chemin vers le fichier PDF.
            
        Returns:
            dict: Résultats de l'analyse.
        """
        st.info("Analyse du PDF en cours... Cette opération peut prendre quelques instants.")
        progress_bar = st.progress(0)
        
        try:
            # Utiliser pdf2image pour convertir le PDF en images
            from pdf2image import convert_from_path
            
            # Sur Streamlit Cloud, nous devons spécifier le chemin vers poppler
            try:
                images = convert_from_path(pdf_path)
            except Exception as e:
                st.error(f"Erreur lors de la conversion du PDF: {str(e)}")
                st.warning("Tentative d'installation de poppler...")
                
                # Afficher une information pour l'installation manuelle de poppler dans requirements.txt
                st.info("Assurez-vous que 'poppler-utils' est installé sur votre serveur Streamlit Cloud. Ajoutez les dépendances dans packages.txt")
                return {"error": "L'installation de poppler est requise pour traiter les PDFs."}
        
            extracted_text = ""
            total_pages = len(images)
            
            for i, image in enumerate(images):
                # Mettre à jour la barre de progression
                progress_bar.progress((i + 0.5) / total_pages)
                
                # Prétraitement de l'image pour améliorer l'OCR
                img_np = np.array(image)
                
                # Conversion en niveaux de gris
                gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
                
                # Appliquer un seuillage pour améliorer la netteté
                _, binary = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
                
                # Effectuer l'OCR
                page_text = pytesseract.image_to_string(binary, config=self.tesseract_config, lang='fra')
                
                # Ajouter le texte extrait
                extracted_text += f"\n\n--- PAGE {i+1} ---\n\n{page_text}"
                
                # Mettre à jour la barre de progression
                progress_bar.progress((i + 1) / total_pages)
        
            # D'abord essayer l'extraction directe
            direct_result = self._extract_charges_data(extracted_text)
            
            # Vérifier si l'extraction directe a donné des résultats satisfaisants
            if self._is_extraction_complete(direct_result):
                st.success("Analyse réussie avec l'extraction directe!")
                return direct_result
            else:
                # Sinon, utiliser GPT-4o-mini
                st.info("Utilisation de l'IA pour compléter l'analyse...")
                return self._analyze_with_gpt(extracted_text, source_type="PDF")
            
        except Exception as e:
            st.error(f"Erreur lors de l'analyse du PDF: {str(e)}")
            return {"error": str(e)}
    
    def analyze_excel(self, excel_path):
        """
        Analyse un fichier Excel.
        
        Args:
            excel_path (str): Chemin vers le fichier Excel.
            
        Returns:
            dict: Résultats de l'analyse.
        """
        st.info("Analyse du fichier Excel en cours...")
        
        try:
            # Lire toutes les feuilles du fichier Excel
            import openpyxl
            excel_data = {}
            xlsx = openpyxl.load_workbook(excel_path, data_only=True)
            
            for sheet_name in xlsx.sheetnames:
                sheet = xlsx[sheet_name]
                
                # Convertir la feuille en texte
                sheet_text = f"Feuille: {sheet_name}\n\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    sheet_text += " | ".join(row_values) + "\n"
                
                excel_data[sheet_name] = sheet_text
            
            # Concaténer toutes les feuilles en un seul texte
            combined_text = "\n\n".join([f"--- {sheet} ---\n{text}" for sheet, text in excel_data.items()])
            
            # D'abord essayer l'extraction directe
            direct_result = self._extract_charges_data(combined_text)
            
            # Vérifier si l'extraction directe a donné des résultats satisfaisants
            if self._is_extraction_complete(direct_result):
                st.success("Analyse réussie avec l'extraction directe!")
                return direct_result
            else:
                # Sinon, utiliser GPT-4o-mini
                st.info("Utilisation de l'IA pour compléter l'analyse...")
                return self._analyze_with_gpt(combined_text, source_type="Excel")
            
        except Exception as e:
            st.error(f"Erreur lors de l'analyse du fichier Excel: {str(e)}")
            return {"error": str(e)}
    
    def analyze_csv(self, csv_path):
        """
        Analyse un fichier CSV.
        
        Args:
            csv_path (str): Chemin vers le fichier CSV.
            
        Returns:
            dict: Résultats de l'analyse.
        """
        st.info("Analyse du fichier CSV en cours...")
        
        try:
            # Essayer de déterminer le délimiteur
            with open(csv_path, 'r', encoding='utf-8', errors='replace') as file:
                sample = file.read(1024)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
            
            # Lire le CSV
            df = pd.read_csv(csv_path, delimiter=delimiter, encoding='utf-8', errors='replace')
            
            # Convertir en texte
            csv_text = df.to_string(index=False)
            
            # D'abord essayer l'extraction directe
            direct_result = self._extract_charges_data(csv_text)
            
            # Vérifier si l'extraction directe a donné des résultats satisfaisants
            if self._is_extraction_complete(direct_result):
                st.success("Analyse réussie avec l'extraction directe!")
                return direct_result
            else:
                # Sinon, utiliser GPT-4o-mini
                st.info("Utilisation de l'IA pour compléter l'analyse...")
                return self._analyze_with_gpt(csv_text, source_type="CSV")
            
        except Exception as e:
            st.error(f"Erreur lors de l'analyse du fichier CSV: {str(e)}")
            return {"error": str(e)}
    
    def _analyze_with_gpt(self, text, source_type):
        """
        Utilise GPT-4o-mini pour analyser le texte extrait.
        
        Args:
            text (str): Texte à analyser.
            source_type (str): Type de la source (PDF, Excel, CSV).
            
        Returns:
            dict: Résultats de l'analyse structurés.
        """
        logger.info(f"Analyse du texte avec GPT-4o-mini (source: {source_type})")
        
        # Limiter la taille du texte si nécessaire
        if len(text) > 15000:
            logger.warning(f"Le texte a été tronqué de {len(text)} à 15000 caractères")
            text = text[:15000] + "...[tronqué]"
        
        # Définir le prompt pour GPT-4o-mini
        prompt = f"""
        Tu es un expert en analyse de tableaux de charges et redditions de charges pour des baux commerciaux.
        Voici un tableau de charges extrait d'un document {source_type}.
        
        {text}
        
        Analyse précisément ce tableau de charges et extrait les informations suivantes au format JSON :
        1. Type de document : détermine s'il s'agit d'une reddition de charges, d'un appel de charges ou d'un budget prévisionnel
        2. Année ou période concernée : extrait la période exacte mentionnée (format JJ/MM/AAAA)
        3. Montant total des charges : montant total en euros
        4. Répartition des charges par poste : extrait tous les postes mentionnés (nettoyage, sécurité, maintenance, etc.) avec leurs montants
        5. Quote-part du locataire : fraction ou pourcentage des charges affectées au locataire (tantièmes)
        6. Montant facturé au locataire : montant final à payer par le locataire
        7. Montant des provisions déjà versées : sommes déjà payées par le locataire
        8. Solde : différence entre les provisions et les charges réelles (créditeur ou débiteur)
        
        Renvoie le JSON avec les clés EXACTEMENT dans ce format standardisé :
        {
          "Type de document": "reddition de charges",
          "Année ou période concernée": "01/01/2022 au 31/12/2022", 
          "Montant total des charges": 28880.45,
          "Répartition des charges par poste": {
            "NETTOYAGE EXTERIEUR": 3903.08,
            "DECHETS SECS": 2198.50
          },
          "Quote-part du locataire": "8565/2092.00 (365 jours)",
          "Montant facturé au locataire": -2499.55,
          "Montant des provisions déjà versées": -31380.00,
          "Solde": -2499.55
        }
        
        IMPORTANT: Réponds uniquement avec le JSON structuré des résultats, sans aucun texte d'introduction ni formatage supplémentaire comme des backticks.
        Ne pas inclure "```json" ni "```" dans ta réponse. Renvoie uniquement le JSON pur.
        """
        
        try:
            # Appeler l'API OpenAI
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Tu es un assistant spécialisé dans l'analyse de documents financiers immobiliers. Tu réponds uniquement en JSON valide sans formatage supplémentaire."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1
            )
            
            # Extraire la réponse
            analysis_text = response.choices[0].message.content
            
            # Tenter d'extraire le JSON avec notre méthode sécurisée
            analysis_result = self._extract_json_safely(analysis_text)
            
            if analysis_result:
                st.success("Analyse réussie avec l'IA!")
                return analysis_result
            else:
                st.error("Impossible d'extraire un JSON valide de la réponse")
                st.text_area("Réponse brute:", analysis_text, height=300, key="raw_response_error_1")
                return {"error": "Format de réponse invalide", "raw_response": analysis_text}
            
        except Exception as e:
            st.error(f"Erreur lors de l'analyse: {str(e)}")
            return {"error": str(e)}
    
    def _extract_json_safely(self, text):
        """
        Tente plusieurs approches pour extraire un JSON valide d'un texte.
        Cette version améliorée traite les cas spécifiques de répartition des charges.
        
        Args:
            text (str): Texte contenant potentiellement du JSON.
            
        Returns:
            dict: Le dictionnaire JSON extrait ou None si l'extraction échoue.
        """
        # Supprimer tous les backticks et mentions "json"
        cleaned_text = text.replace('```json', '').replace('```', '').replace('json', '').strip()
        
        # Définir une fonction pour corriger le format de la répartition des charges
        def traiter_repartition_charges(json_obj):
            """Corrige le format de la répartition des charges si nécessaire."""
            if not isinstance(json_obj, dict):
                return json_obj
                
            if "Répartition des charges par poste" in json_obj and isinstance(json_obj["Répartition des charges par poste"], str):
                repartition_str = json_obj["Répartition des charges par poste"]
                repartition_dict = {}
                
                # Si c'est déjà un format JSON valide, essayer de le parser
                if repartition_str.startswith('{') and repartition_str.endswith('}'):
                    try:
                        import json
                        repartition_dict = json.loads(repartition_str)
                        json_obj["Répartition des charges par poste"] = repartition_dict
                        return json_obj
                    except:
                        # Si le parsing échoue, continuer avec la méthode manuelle
                        repartition_str = repartition_str[1:-1]  # Enlever les accolades
                    
                # Format "POSTE1: VALEUR1, POSTE2: VALEUR2"
                try:
                    pairs = repartition_str.split(',')
                    for pair in pairs:
                        if ':' in pair:
                            key, value = pair.split(':', 1)
                            key = key.strip().strip('"\'')  # Enlever guillemets et espaces
                            value = value.strip().strip('"\'')  # Enlever guillemets et espaces
                            
                            # Convertir la valeur en nombre si possible
                            try:
                                value = float(value.replace(' ', ''))
                            except:
                                pass  # Garder comme chaîne si la conversion échoue
                                
                            repartition_dict[key] = value
                    
                    json_obj["Répartition des charges par poste"] = repartition_dict
                except Exception as e:
                    # Logging de l'erreur pour debug
                    import logging
                    logging.error(f"Erreur lors du traitement de la répartition: {str(e)}")
                    
            return json_obj
        
        # Prétraitement pour convertir les nombres avec format français (virgule) en format anglo-saxon (point)
        def convert_numbers_in_json(json_text):
            # Fonction pour remplacer les nombres avec virgule par des nombres avec point dans un JSON
            # Pattern qui cherche des nombres (ex: 3903,08) suivis d'une virgule ou d'un '}' ou d'un ']'
            pattern = r'(\d+),(\d+)(?=\s*[,}\]])'
            return re.sub(pattern, r'\1.\2', json_text)
        
        # Appliquer la conversion des nombres
        cleaned_text = convert_numbers_in_json(cleaned_text)
        
        # Approche 1: Essayer de charger directement le texte nettoyé
        try:
            result = json.loads(cleaned_text)
            return traiter_repartition_charges(result)
        except json.JSONDecodeError:
            pass
        
        # Approche 2: Chercher les délimiteurs {} les plus externes
        try:
            start_idx = cleaned_text.find('{')
            end_idx = cleaned_text.rfind('}') + 1
            
            if start_idx >= 0 and end_idx > start_idx:
                json_str = cleaned_text[start_idx:end_idx]
                # Convertir les nombres avec virgule
                json_str = convert_numbers_in_json(json_str)
                result = json.loads(json_str)
                return traiter_repartition_charges(result)
        except json.JSONDecodeError:
            # Utiliser une clé unique pour chaque affichage de texte
            st.text_area("Texte JSON problématique:", cleaned_text, height=200, key="json_error_text_1")
            pass
        
        # Approche 3: Corriger les problèmes courants dans les JSON générés par l'IA
        try:
            # Remplacer les guillemets simples par des guillemets doubles
            fixed_text = cleaned_text.replace("'", '"')
            
            # Ajouter des guillemets aux clés non quotées (pattern: key: value)
            fixed_text = re.sub(r'(\w+):\s*', r'"\1": ', fixed_text)
            
            # Corriger le format de la répartition des charges par poste si présente
            fixed_text = re.sub(r'"Répartition des charges par poste":\s*"([^"]*)"', 
                               lambda m: f'"Répartition des charges par poste": {{{m.group(1)}}}', 
                               fixed_text)
            
            # Convertir les nombres avec virgule en nombres avec point
            fixed_text = convert_numbers_in_json(fixed_text)
            
            # Corriger les guillemets dans les valeurs de répartition
            fixed_text = re.sub(r'"([^"]+)":\s*"(\d+)[,.](\d+)"', r'"\1": \2.\3', fixed_text)
            
            result = json.loads(fixed_text)
            return traiter_repartition_charges(result)
        except json.JSONDecodeError:
            pass
        
        # Approche 4: Transformer manuellement le format des valeurs numériques
        try:
            # Remplacer les valeurs numériques qui sont entre guillemets
            number_pattern = r'"(\d+[,.]\d+)"'
            
            def replace_quoted_numbers(match):
                # Remplacer la virgule par un point et enlever les guillemets
                num_str = match.group(1).replace(',', '.')
                return num_str  # Sans guillemets
            
            numeric_fixed_text = re.sub(number_pattern, replace_quoted_numbers, cleaned_text)
            result = json.loads(numeric_fixed_text)
            return traiter_repartition_charges(result)
        except json.JSONDecodeError:
            pass
        
        # Approche 5: Utiliser une expression régulière pour extraire le JSON
        try:
            json_pattern = r'(\{.*\})'
            match = re.search(json_pattern, cleaned_text, re.DOTALL)
            if match:
                json_str = match.group(1)
                # Convertir les nombres avec virgule
                json_str = convert_numbers_in_json(json_str)
                result = json.loads(json_str)
                return traiter_repartition_charges(result)
        except json.JSONDecodeError:
            pass
        
        # Approche 6: Tentative de correction manuelle de formats JSON courants
        try:
            # Correction de certaines erreurs de formatage fréquentes
            corrected_text = cleaned_text
            # Remplacer les virgules à la fin des objets
            corrected_text = re.sub(r',\s*}', '}', corrected_text)
            # Ajouter des guillemets aux clés non quotées
            corrected_text = re.sub(r'(\w+):', r'"\1":', corrected_text)
            # Convertir tous les nombres avec virgule en format point
            corrected_text = convert_numbers_in_json(corrected_text)
            
            # Tentative supplémentaire pour corriger le format de nombres dans les valeurs
            # Remplacer "POSTE": "3903,08" par "POSTE": 3903.08 (sans guillemets)
            corrected_text = re.sub(r'"([^"]+)"\s*:\s*"(\d+)[,.](\d+)"', r'"\1": \2.\3', corrected_text)
            
            result = json.loads(corrected_text)
            return traiter_repartition_charges(result)
        except json.JSONDecodeError:
            # Afficher le texte corrigé en cas d'échec avec une clé unique
            st.text_area("Tentative de correction JSON échouée:", corrected_text, height=200, key="json_error_text_2")
            pass
        
        # Approche 7: Extraction manuelle pour le document spécifique si toutes les autres méthodes échouent
        try:
            # Cette approche est un dernier recours lorsque nous sommes sûrs du format du document
            # Elle est utile pour les documents qui suivent toujours le même modèle
            manual_json = self._extract_reddition_charges_manually(text)
            if manual_json:
                return manual_json
        except Exception as e:
            st.error(f"Erreur lors de l'extraction manuelle: {str(e)}")
        
        # Si toutes les tentatives échouent
        return None

    def _extract_reddition_charges_manually(self, text):
        """
        Extraction manuelle des informations d'une reddition de charges
        en fonction du format du document.
        
        Args:
            text (str): Le texte extrait du document.
            
        Returns:
            dict: Les informations extraites ou None si l'extraction échoue.
        """
        # Détecter le type de document
        type_document = "Reddition de charges" if "REDDITION" in text or "reddition" in text.lower() else "Non spécifié"
        
        # Extraire la période concernée
        periode = "Non spécifiée"
        periode_match = re.search(r'[Pp]ériode du (\d{2}/\d{2}/\d{4}) au (\d{2}/\d{2}/\d{4})', text)
        if periode_match:
            periode = f"{periode_match.group(1)} au {periode_match.group(2)}"
        
        # Extraire le montant total des charges
        montant_total = None
        # Plusieurs patterns possibles
        for pattern in [
            r'Total charges\s+(\d+[\s\xa0]*\d*[,.]\d{2})',
            r'Charges\s+(\d+[\s\xa0]*\d*[,.]\d{2})',
            r'Total Chap[\.\s]+\d+[\s\-]+[A-Z\s]+\s+(\d+[\s\xa0]*\d*[,.]\d{2})'
        ]:
            match = re.search(pattern, text)
            if match:
                valeur = match.group(1).replace(' ', '').replace('\xa0', '').replace(',', '.')
                try:
                    montant_total = float(valeur)
                    break
                except ValueError:
                    continue
        
        # Extraire les provisions versées
        provisions = None
        provisions_match = re.search(r'Provisions\s+(-?\d+[\s\xa0]*\d*[,.]\d{2})', text)
        if provisions_match:
            valeur = provisions_match.group(1).replace(' ', '').replace('\xa0', '').replace(',', '.')
            try:
                provisions = float(valeur)
            except ValueError:
                provisions = "Non spécifié"
        
        # Extraire le solde
        solde = None
        solde_match = re.search(r'Solde\s+(-?\d+[\s\xa0]*\d*[,.]\d{2})', text)
        if solde_match:
            valeur = solde_match.group(1).replace(' ', '').replace('\xa0', '').replace(',', '.')
            try:
                solde = float(valeur)
            except ValueError:
                solde = "Non spécifié"
        
        # Extraire la quote-part
        quote_part = "Non spécifiée"
        if "Tantièmes" in text:
            # Format: "8565    2092.00    365"
            tantiemes_match = re.search(r'(\d+)\s+(\d+\.\d{2})\s+365', text)
            if tantiemes_match:
                quote_part = f"{tantiemes_match.group(1)}/{tantiemes_match.group(2)} (365 jours)"
        
        # Extraire la répartition des charges par poste
        repartition_charges = {}
        
        # Ce patron recherche les lignes comme "NETTOYAGE EXTERIEUR    15 979,86 €    8565    2092.00    365    3 903,08"
        # [poste]    [montant total]    [tantièmes]    [tantièmes]    [jours]    [montant locataire]
        charges_lines = re.findall(r'(\d{2}\s*/\s*\d{2}|\d{2}/\d{2})\s+([A-ZÀ-Ú\s/&.]+)\s+(\d+[\s\xa0]*\d*[,.]\d{2})', text)
        
        for match in charges_lines:
            code_poste = match[0].strip()
            poste = match[1].strip()
            
            # Chercher le montant spécifique du locataire (généralement à la fin de la ligne)
            poste_line = re.search(f"{re.escape(poste)}.*?(\d+[\s\xa0]*\d*[,.]\d{{2}})\s*$", text, re.MULTILINE)
            if poste_line:
                montant_str = poste_line.group(1).replace(' ', '').replace('\xa0', '').replace(',', '.')
                try:
                    montant = float(montant_str)
                    repartition_charges[poste] = montant
                except ValueError:
                    continue
        
        # Si aucune répartition n'a été trouvée, essayer une approche plus générique
        if not repartition_charges:
            # Chercher des lignes contenant des postes et des montants
            charge_lines = re.findall(r'([A-ZÀ-Ú\s/&.]+)\s+(\d+[\s\xa0]*\d*[,.]\d{2})\s*€?\s*', text)
            for poste, montant_str in charge_lines:
                poste = poste.strip()
                if poste and len(poste) > 3:  # Éviter les faux positifs
                    try:
                        montant = float(montant_str.replace(' ', '').replace('\xa0', '').replace(',', '.'))
                        repartition_charges[poste] = montant
                    except ValueError:
                        continue
        
        # Construire le résultat final
        resultat = {
            "Type de document": type_document,
            "Année ou période concernée": periode,
            "Montant total des charges": montant_total if montant_total is not None else "Non spécifié",
            "Répartition des charges par poste": repartition_charges,
            "Quote-part du locataire": quote_part,
            "Montant facturé au locataire": solde if solde is not None else "Non spécifié",
            "Montant des provisions déjà versées": provisions if provisions is not None else "Non spécifié",
            "Solde": solde if solde is not None else "Non spécifié"
        }
        
        return resultat
    
    def _extract_charges_data(self, text):
    """
    Méthode d'extraction directe des données de charges sans recourir à l'API.
    Version améliorée qui traite les différents formats de documents de charges.
    
    Args:
        text (str): Le texte extrait du document.
        
    Returns:
        dict: Les informations extraites.
    """
    import re
    
    # Normaliser les espaces et les caractères spéciaux
    text = text.replace('\xa0', ' ')
    
    # Détecter le type de document
    type_document = "Non spécifié"
    if "REDDITION" in text or "reddition" in text.lower():
        type_document = "Reddition de charges"
    elif "APPEL DE CHARGES" in text or "appel de charges" in text.lower():
        type_document = "Appel de charges"
    elif "BUDGET PRÉVISIONNEL" in text or "budget prévisionnel" in text.lower():
        type_document = "Budget prévisionnel"
        
    # Extraire la période concernée
    periode = "Non spécifiée"
    periode_patterns = [
        r'[Pp]ériode du (\d{2}/\d{2}/\d{4}) au (\d{2}/\d{2}/\d{4})',
        r'[Pp]ériode\s*:\s*(\d{2}/\d{2}/\d{4}) au (\d{2}/\d{2}/\d{4})',
        r'du (\d{2}/\d{2}/\d{4}) au (\d{2}/\d{2}/\d{4})'
    ]
    
    for pattern in periode_patterns:
        periode_match = re.search(pattern, text)
        if periode_match:
            periode = f"{periode_match.group(1)} au {periode_match.group(2)}"
            break
            
    # Extraire le montant total des charges
    montant_total = "Non spécifié"
    # Plusieurs patterns possibles selon le format du document
    patterns_montant = [
        r'Total charges\s+([0-9\s]+[,.]\d{2})',
        r'Charges\s+([0-9\s]+[,.]\d{2}\s*€?)',
        r'REDDITION CHARGES[^\n]+\s+([0-9\s]+[,.]\d{2}\s*€?)',
        r'Total clé.*?CHARGES COMMUNES\s+([0-9\s]+[,.]\d{2})',
        r'Total Chap\. \d+ [A-Z\s]+\s+([0-9\s]+[,.]\d{2})'
    ]
    
    for pattern in patterns_montant:
        match = re.search(pattern, text)
        if match:
            montant_str = match.group(1).replace(' ', '').replace(',', '.').replace('€', '')
            try:
                montant_total = float(montant_str)
            except ValueError:
                montant_total = montant_str
            break
            
    # Extraire le montant des provisions versées
    provisions = "Non spécifié"
    provisions_patterns = [
        r'Provisions\s+(-?[0-9\s]+[,.]\d{2}\s*€?)',
        r'RBT PROV\. CHARGES.*?(-[0-9\s]+[,.]\d{2})',
        r'Provisions déjà versées\s+(-?[0-9\s]+[,.]\d{2})'
    ]
    
    for pattern in provisions_patterns:
        provisions_match = re.search(pattern, text)
        if provisions_match:
            provisions_str = provisions_match.group(1).replace(' ', '').replace(',', '.').replace('€', '')
            try:
                provisions = float(provisions_str)
            except ValueError:
                provisions = provisions_str
            break
            
    # Extraire le solde
    solde = "Non spécifié"
    solde_patterns = [
        r'Solde\s+(-?[0-9\s]+[,.]\d{2}\s*€?)',
        r'SOLDE\s+(-?[0-9\s]+[,.]\d{2})',
        r'Total H\.T\.\s*:\s*(-?[0-9\s]+[,.]\d{2})\s*€?',
        r'Total\s*:\s*(-?[0-9\s]+[,.]\d{2})\s*€?'
    ]
    
    for pattern in solde_patterns:
        solde_match = re.search(pattern, text)
        if solde_match:
            solde_str = solde_match.group(1).replace(' ', '').replace(',', '.').replace('€', '')
            try:
                solde = float(solde_str)
            except ValueError:
                solde = solde_str
            break
            
    # Extraire la quote-part du locataire
    quote_part = "Non spécifiée"
    
    # Différents formats possibles
    if "Tantièmes" in text and ("particuliers" in text or "Quote-part" in text):
        # Format tableau avec tantièmes
        # Rechercher les tantièmes spécifiques au locataire
        tantiemes_pattern = r'(\d+)\s+(\d+\.\d{2})\s+365\s+'
        tantiemes_match = re.search(tantiemes_pattern, text)
        if tantiemes_match:
            quote_part = f"{tantiemes_match.group(1)}/{tantiemes_match.group(2)} (365 jours)"
    elif "Quote-part" in text or "quote-part" in text.lower():
        # Autres formats de quote-part
        quote_part_pattern = r'[Qq]uote-part\s*:?\s*(\d+[,.]\d+\s*%)'
        quote_part_match = re.search(quote_part_pattern, text)
        if quote_part_match:
            quote_part = quote_part_match.group(1)
            
    # Extraire la répartition des charges par poste
    repartition_charges = {}
    
    # Plusieurs patterns possibles selon le format du document
    
    # 1. Format avec codes de postes (ex: 01/01 NETTOYAGE EXTERIEUR)
    charges_pattern_1 = r'(\d{2}\s*/\s*\d{2}|\d{2}/\d{2})\s+([A-ZÀ-Ú\s/&.]+)(?:\s+)(\d+[\s\xa0]*\d*[,.]\d{2}\s*€?)(?:.*?)(\d+[\s\xa0]*\d*[,.]\d{2})'
    
    for match in re.finditer(charges_pattern_1, text):
        poste = match.group(2).strip()
        montant_str = match.group(4).replace(' ', '').replace('\xa0', '').replace(',', '.')
        try:
            montant = float(montant_str)
            repartition_charges[poste] = montant
        except ValueError:
            continue
    
    # 2. Format sans code, directement le nom du poste et le montant
    if not repartition_charges:
        charges_pattern_2 = r'([A-ZÀ-Ú\s/&.]{5,})\s+(\d+[\s\xa0]*\d*[,.]\d{2})\s*€?\s*
        
        for match in re.finditer(charges_pattern_2, text, re.MULTILINE):
            poste = match.group(1).strip()
            # Éviter les faux positifs (lignes de total, etc.)
            if poste in ["TOTAL", "SOLDE", "PROVISIONS", "TOTAL CHARGES"]:
                continue
                
            montant_str = match.group(2).replace(' ', '').replace('\xa0', '').replace(',', '.')
            try:
                montant = float(montant_str)
                repartition_charges[poste] = montant
            except ValueError:
                continue
    
    # 3. Format tableau avec tantièmes (comme dans le document fourni)
    if not repartition_charges:
        # Format: "POSTE    MONTANT_TOTAL    8565    2092.00    365    MONTANT_LOCATAIRE"
        charges_pattern_3 = r'([A-ZÀ-Ú\s/&.]{5,})\s+\d+[\s\xa0]*\d*[,.]\d{2}\s*€?\s+\d+\s+[\d\.]+\s+\d+\s+(\d+[\s\xa0]*\d*[,.]\d{2})'
        
        for match in re.finditer(charges_pattern_3, text):
            poste = match.group(1).strip()
            montant_str = match.group(2).replace(' ', '').replace('\xa0', '').replace(',', '.')
            try:
                montant = float(montant_str)
                repartition_charges[poste] = montant
            except ValueError:
                continue
    
    # Extraire le montant facturé au locataire
    montant_facture = "Non spécifié"
    
    # Pour une reddition, le montant facturé est généralement le solde
    if type_document == "Reddition de charges" and solde != "Non spécifié":
        montant_facture = solde
    else:
        # Autres formats possibles
        facture_patterns = [
            r'[Mm]ontant (?:à payer|facturé)\s*:?\s*(-?[0-9\s]+[,.]\d{2}\s*€?)',
            r'Total T\.T\.C\.\s*:?\s*(-?[0-9\s]+[,.]\d{2}\s*€?)'
        ]
        
        for pattern in facture_patterns:
            facture_match = re.search(pattern, text)
            if facture_match:
                montant_str = facture_match.group(1).replace(' ', '').replace(',', '.').replace('€', '')
                try:
                    montant_facture = float(montant_str)
                except ValueError:
                    montant_facture = montant_str
                break
    
    # Construire le résultat
    result = {
        "Type de document": type_document,
        "Année ou période concernée": periode,
        "Montant total des charges": montant_total,
        "Répartition des charges par poste": repartition_charges,
        "Quote-part du locataire": quote_part,
        "Montant facturé au locataire": montant_facture,
        "Montant des provisions déjà versées": provisions,
        "Solde": solde
    }
    
    return result
